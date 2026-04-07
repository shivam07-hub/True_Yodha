"""
manual_tag_exporter.py
Human-in-the-loop alternative to groq_tagger.py.

Step 1 of 2: Export untagged jobs to Excel for manual LLM tagging.

Usage:
    python3 -m app.services.manual_tag_exporter
    python3 -m app.services.manual_tag_exporter --batch-size 50
    python3 -m app.services.manual_tag_exporter --batch-size 100 --batch-num 2

Output:
    Market Data/Job_Scrapers/All_CSV_Outputs/Manual_Tag/
        MANUAL_TAG_BATCH_<N>_<date>.xlsx   ← paste into LLM
        PROMPT_TEMPLATE.txt                ← copy-paste this as your LLM prompt

Workflow:
    1. Run this script → get Excel + prompt template
    2. Open PROMPT_TEMPLATE.txt, copy the prompt
    3. Upload / paste the Excel (or its contents) into Claude / ChatGPT
    4. Paste the LLM response back into the Excel — fill in the 5 tag columns
    5. Save the filled Excel
    6. Run manual_tag_importer.py to merge tags into cache
    7. Repeat for remaining batches
    8. Run csv_importer.py to push to Supabase
"""

import argparse
import json
import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

PROJECT_ROOT = Path(__file__).resolve().parents[3]
TAXONOMY_JSON = PROJECT_ROOT / "skill_taxonomy mapping" / "taxonomy.json"
MASTER_OUTPUT_DIR = (
    PROJECT_ROOT / "Market Data" / "Job_Scrapers" / "All_CSV_Outputs" / "Master_Output"
)
MANUAL_TAG_DIR = (
    PROJECT_ROOT / "Market Data" / "Job_Scrapers" / "All_CSV_Outputs" / "Manual_Tag"
)
CACHE_FILE = PROJECT_ROOT / "database" / "skill_tag_cache.json"

DEFAULT_BATCH_SIZE = 50

# Columns the user / LLM must fill in (highlighted yellow)
TAG_COLUMNS = [
    "required_skills",
    "preferred_skills",
    "unknown_skills",
    "min_years_experience",
    "min_qualification",
]


# ── Taxonomy ──────────────────────────────────────────────────────────────────

def load_taxonomy_keys(json_path: Path) -> list[str]:
    with open(json_path, encoding="utf-8") as f:
        taxonomy = json.load(f)
    keys: set[str] = set()
    for _cat, skills in taxonomy["categories"].items():
        for skill_key in skills:
            keys.add(skill_key)
    return sorted(keys)


# ── Jobs ──────────────────────────────────────────────────────────────────────

def load_cache() -> dict[str, dict]:
    if CACHE_FILE.exists():
        with open(CACHE_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}


def read_all_jobs(master_files: list[Path]) -> list[dict]:
    seen: set[str] = set()
    jobs: list[dict] = []
    for xlsx_path in master_files:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(h) for h in next(rows)]
        for row in rows:
            col = {h: row[i] for i, h in enumerate(headers)}
            jid = str(col.get("job_id", "")).strip()
            title = str(col.get("title", "")).strip()
            if not jid or not title or title == "See full role description":
                continue
            if jid in seen:
                continue
            seen.add(jid)
            jobs.append({
                "job_id": jid,
                "title": title,
                "company": str(col.get("company_name") or "").strip() or None,
                "raw_jd_text": str(col.get("raw_jd_text") or "").strip() or None,
            })
        wb.close()
    return jobs


# ── Excel writer ──────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TAG_FILL    = PatternFill("solid", fgColor="FFF2CC")  # yellow — LLM fills these
_TAG_FONT    = Font(color="7F6000")

_COLUMNS = ["job_id", "title", "company", "description_preview"] + TAG_COLUMNS


def write_export_xlsx(jobs: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tag These Jobs"

    # Header row
    for col_idx, col_name in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    tag_col_indices = {name: _COLUMNS.index(name) + 1 for name in TAG_COLUMNS}

    for row_idx, job in enumerate(jobs, 2):
        preview = (job.get("raw_jd_text") or "")[:2000]
        ws.cell(row=row_idx, column=1, value=job["job_id"])
        ws.cell(row=row_idx, column=2, value=job["title"])
        ws.cell(row=row_idx, column=3, value=job.get("company") or "")
        ws.cell(row=row_idx, column=4, value=preview)
        for col_name in TAG_COLUMNS:
            cell = ws.cell(row=row_idx, column=tag_col_indices[col_name], value="")
            cell.fill = _TAG_FILL
            cell.font = _TAG_FONT

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 80
    for col_name in TAG_COLUMNS:
        letter = openpyxl.utils.get_column_letter(tag_col_indices[col_name])
        ws.column_dimensions[letter].width = 35

    wb.save(output_path)


def write_prompt_template(taxonomy_keys: list[str], output_path: Path) -> None:
    taxonomy_str = "\n".join(f"  - {k}" for k in taxonomy_keys)
    template = f"""=== MIRROR — JOB SKILL TAGGING PROMPT ===

Paste or attach the Excel sheet, then send this prompt:

---

You are an expert job analyst. For each row in the attached Excel, fill in the 5 yellow columns:

1. required_skills
   Comma-separated taxonomy keys for ESSENTIAL skills.
   Use ONLY keys from this list (exact match, lowercase):
{taxonomy_str}

2. preferred_skills
   Comma-separated taxonomy keys for NICE-TO-HAVE skills.
   Same taxonomy list. A skill cannot appear in both required and preferred.

3. unknown_skills
   Comma-separated skill names clearly required but NOT in the taxonomy.
   Raw lowercase names, max 5.

4. min_years_experience
   Integer (e.g. 3) or leave blank if not specified.

5. min_qualification
   One of: High School | Diploma | Bachelor's | Master's | MBA | PhD | Any

Return the SAME Excel with ONLY those 5 columns filled in.
Do not modify any other columns. Do not add rows.
Return one row per job, same order as input.

---
"""
    output_path.write_text(template, encoding="utf-8")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Export untagged jobs to Excel for manual LLM tagging")
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE,
                        help=f"Jobs per Excel file (default: {DEFAULT_BATCH_SIZE})")
    parser.add_argument("--batch-num", type=int, default=1,
                        help="Which batch to export (default: 1 = first untagged batch)")
    args = parser.parse_args()

    master_files = sorted(MASTER_OUTPUT_DIR.glob("ALL_JOBS_NORMALIZED_*.xlsx"))
    if not master_files:
        print(f"ERROR: No master xlsx files in {MASTER_OUTPUT_DIR}")
        sys.exit(1)

    print("Loading taxonomy...")
    taxonomy_keys = load_taxonomy_keys(TAXONOMY_JSON)
    print(f"  {len(taxonomy_keys)} skills in taxonomy")

    print("Reading jobs...")
    all_jobs = read_all_jobs(master_files)
    print(f"  {len(all_jobs)} unique jobs loaded")

    cache = load_cache()
    untagged = [j for j in all_jobs if j["job_id"] not in cache]
    print(f"  {len(cache)} already cached, {len(untagged)} untagged")

    if not untagged:
        print("All jobs already tagged. Nothing to export.")
        sys.exit(0)

    batch_size = args.batch_size
    batch_num = args.batch_num
    total_batches = (len(untagged) + batch_size - 1) // batch_size

    if batch_num < 1 or batch_num > total_batches:
        print(f"ERROR: --batch-num must be between 1 and {total_batches}")
        sys.exit(1)

    start = (batch_num - 1) * batch_size
    batch = untagged[start: start + batch_size]

    today = date.today().strftime("%Y%m%d")
    xlsx_path = MANUAL_TAG_DIR / f"MANUAL_TAG_BATCH_{batch_num:03d}_{today}.xlsx"
    prompt_path = MANUAL_TAG_DIR / "PROMPT_TEMPLATE.txt"

    print(f"\nExporting batch {batch_num}/{total_batches} ({len(batch)} jobs)...")
    write_export_xlsx(batch, xlsx_path)
    write_prompt_template(taxonomy_keys, prompt_path)

    print(f"\nDone.")
    print(f"  Excel:   {xlsx_path}")
    print(f"  Prompt:  {prompt_path}")
    print(f"\nNext steps:")
    print(f"  1. Open PROMPT_TEMPLATE.txt — copy the prompt")
    print(f"  2. Upload the Excel to Claude / ChatGPT with the prompt")
    print(f"  3. Fill the yellow columns with the LLM output")
    print(f"  4. Save the Excel")
    print(f"  5. Run: python3 -m app.services.manual_tag_importer --file \"{xlsx_path}\"")
    if batch_num < total_batches:
        print(f"  6. Export next batch: python3 -m app.services.manual_tag_exporter --batch-num {batch_num + 1}")


if __name__ == "__main__":
    main()
