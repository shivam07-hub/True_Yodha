"""
groq_tagger.py
Stage 1 of the two-stage import pipeline.

Reads raw master xlsx → preprocesses → tags skills with Groq → writes enhanced xlsx.

Run from project root:
    python3 backend/app/services/groq_tagger.py

Output:
    Market Data/Job_Scrapers/All_CSV_Outputs/Enhanced/ENHANCED_JOBS_<YYYYMMDD>.xlsx

The enhanced xlsx is human-reviewable before anything reaches the database.
Run csv_importer.py next to push the reviewed output to Supabase.

Cache:
    Groq results are cached in database/skill_tag_cache.json.
    Re-runs skip already-tagged jobs — safe to resume after interruption.
"""

import json
import os
import sys
from datetime import date
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

from app.services.preprocessor import preprocess

# ── Paths ─────────────────────────────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).resolve().parents[3]
TAXONOMY_JSON = PROJECT_ROOT / "skill_taxonomy mapping" / "taxonomy.json"
MASTER_OUTPUT_DIR = (
    PROJECT_ROOT / "Market Data" / "Job_Scrapers" / "All_CSV_Outputs" / "Master_Output"
)
ENHANCED_OUTPUT_DIR = (
    PROJECT_ROOT / "Market Data" / "Job_Scrapers" / "All_CSV_Outputs" / "Enhanced"
)
ENV_FILE = PROJECT_ROOT / "backend" / ".env"

# ── Taxonomy ──────────────────────────────────────────────────────────────────

def load_taxonomy(json_path: Path) -> dict[str, str]:
    """Returns alias_map: alias_lower → taxonomy_key."""
    with open(json_path, encoding="utf-8") as f:
        taxonomy = json.load(f)
    alias_map: dict[str, str] = {}
    for _category, skills in taxonomy["categories"].items():
        for skill_key, skill_info in skills.items():
            aliases = skill_info["aliases"]
            alias_map[skill_key.lower()] = skill_key
            alias_map[aliases["display_name"].lower()] = skill_key
            for tech in aliases.get("technologies", []):
                alias_map[tech.lower()] = skill_key
    return alias_map

# ── Read raw xlsx ─────────────────────────────────────────────────────────────

def read_all_jobs(master_files: list[Path]) -> list[dict]:
    """Read all master xlsx files. Deduplicates by job_id (exact match).

    Preserves ALL canonical fields from the master file so nothing is silently
    dropped before the enhanced xlsx is written.  The `groq_*` skill columns are
    appended later in write_enhanced_xlsx(); they do not replace anything here.
    """
    seen_ids: set[str] = set()
    jobs: list[dict] = []

    for xlsx_path in master_files:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(h) for h in next(rows)]

        for row in rows:
            col = {h: row[i] for i, h in enumerate(headers)}
            external_id = str(col.get("job_id", "")).strip()
            title = str(col.get("title", "")).strip()

            if not external_id or not title or title == "See full role description":
                continue
            if external_id in seen_ids:
                continue

            seen_ids.add(external_id)
            jobs.append({
                # ── Identity ──────────────────────────────────────────────────
                "job_id": external_id,
                "title": title,
                "company_name": str(col.get("company_name") or "").strip() or None,
                "job_url": str(col.get("job_url") or "").strip() or None,
                "business_unit": col.get("business_unit"),
                "source_platform": col.get("source_platform"),
                # ── Content ───────────────────────────────────────────────────
                "raw_jd_text": str(col.get("raw_jd_text") or "").strip() or None,
                # Pre-existing regex-extracted skills from orchestrator (may be empty)
                "skills_required_raw": str(col.get("skills_required") or ""),
                "skills_preferred_raw": str(col.get("skills_preferred") or ""),
                # ── Structured fields ─────────────────────────────────────────
                "seniority_level": col.get("seniority_level"),
                "location_city": col.get("location_city"),
                "location_country": col.get("location_country"),
                "work_mode": col.get("work_mode"),
                "employment_type": col.get("employment_type"),
                "degree_required": col.get("degree_required"),
                "degree_preferred_field": col.get("degree_preferred_field"),
                "industry": col.get("industry"),
                "min_years_experience": col.get("min_years_experience"),
                "max_years_experience": col.get("max_years_experience"),
                "salary_min": col.get("salary_min"),
                "salary_max": col.get("salary_max"),
                "salary_currency": col.get("salary_currency"),
                "date_posted": col.get("date_posted"),
                "is_active": bool(col.get("is_active", True)),
                # ── Provenance ────────────────────────────────────────────────
                "source_file": xlsx_path.name,
            })

        wb.close()

    return jobs

# ── Write enhanced xlsx ───────────────────────────────────────────────────────

# All canonical fields carried through from the master file, plus groq_* enrichment
# columns appended at the end. Column order matches csv_importer.py expectations.
_ENHANCED_HEADERS = [
    # ── Identity (canonical) ──────────────────────────────────────────────────
    "job_id", "title", "company_name", "job_url",
    "business_unit", "source_platform",
    # ── Location & work (canonical) ───────────────────────────────────────────
    "location_city", "location_country", "work_mode", "employment_type",
    # ── Classification (canonical) ────────────────────────────────────────────
    "seniority_level", "industry",
    "degree_required", "degree_preferred_field",
    # ── Experience (canonical) ────────────────────────────────────────────────
    "min_years_experience", "max_years_experience",
    # ── Compensation (canonical) ──────────────────────────────────────────────
    "salary_min", "salary_max", "salary_currency",
    # ── Dates & flags (canonical) ─────────────────────────────────────────────
    "date_posted", "is_active",
    # ── Content (canonical) ───────────────────────────────────────────────────
    "raw_jd_text",
    "skills_required_raw",    # regex-extracted at scrape time (may be empty)
    "skills_preferred_raw",
    # ── Skill tagging enrichment (groq_* prefix = LM skill tagger output) ────
    "groq_required_skills",
    "groq_preferred_skills",
    "groq_unknown_skills",
    "groq_min_years_experience",
    "groq_min_qualification",
    # ── JD field enrichment (jd_* prefix = structured extraction from raw JD) ─
    "jd_seniority_level",
    "jd_work_mode",
    "jd_employment_type",
    "jd_degree_required",
    "jd_degree_field",
    "jd_industry",
    "jd_min_years_exp",
    "jd_max_years_exp",
    "jd_location_city",
    # ── Provenance ────────────────────────────────────────────────────────────
    "source_file",
]


def write_enhanced_xlsx(
    jobs: list[dict],
    tag_cache: dict[str, dict],
    output_path: Path,
    enrichment_cache: dict[str, dict] | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enhanced Jobs"
    ws.append(_ENHANCED_HEADERS)

    enrichment_cache = enrichment_cache or {}

    for job in jobs:
        jid  = str(job["job_id"])
        tags = tag_cache.get(jid, {})
        enr  = enrichment_cache.get(jid, {})
        ws.append([
            # ── Identity ──────────────────────────────────────────────────────
            job.get("job_id"),
            job.get("title"),
            job.get("company_name"),
            job.get("job_url"),
            job.get("business_unit"),
            job.get("source_platform"),
            # ── Location & work ───────────────────────────────────────────────
            job.get("location_city"),
            job.get("location_country"),
            job.get("work_mode"),
            job.get("employment_type"),
            # ── Classification ────────────────────────────────────────────────
            job.get("seniority_level"),
            job.get("industry"),
            job.get("degree_required"),
            job.get("degree_preferred_field"),
            # ── Experience ────────────────────────────────────────────────────
            job.get("min_years_experience"),
            job.get("max_years_experience"),
            # ── Compensation ──────────────────────────────────────────────────
            job.get("salary_min"),
            job.get("salary_max"),
            job.get("salary_currency"),
            # ── Dates & flags ─────────────────────────────────────────────────
            str(job.get("date_posted") or ""),
            job.get("is_active"),
            # ── Content ───────────────────────────────────────────────────────
            job.get("raw_jd_text"),
            job.get("skills_required_raw", ""),
            job.get("skills_preferred_raw", ""),
            # ── Skill tagging enrichment ───────────────────────────────────────
            ", ".join(tags.get("required_skills", [])),
            ", ".join(tags.get("preferred_skills", [])),
            ", ".join(tags.get("unknown_skills", [])),
            tags.get("min_years_experience"),
            tags.get("min_qualification", "Any"),
            # ── JD field enrichment ───────────────────────────────────────────
            enr.get("jd_seniority_level"),
            enr.get("jd_work_mode"),
            enr.get("jd_employment_type"),
            enr.get("jd_degree_required"),
            enr.get("jd_degree_field"),
            enr.get("jd_industry"),
            enr.get("jd_min_years_exp"),
            enr.get("jd_max_years_exp"),
            enr.get("jd_location_city"),
            # ── Provenance ────────────────────────────────────────────────────
            job.get("source_file"),
        ])

    wb.save(output_path)

# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    load_dotenv(ENV_FILE)
    groq_api_key = os.getenv("GROQ_API_KEY", "")
    lm_studio_tagger_model = os.getenv("LM_STUDIO_TAGGER_MODEL", "")

    if not groq_api_key and not lm_studio_tagger_model:
        print("ERROR: Set at least one of GROQ_API_KEY or LM_STUDIO_TAGGER_MODEL in backend/.env")
        sys.exit(1)

    all_master_files = sorted(MASTER_OUTPUT_DIR.glob("ALL_JOBS_NORMALIZED_*.xlsx"), key=lambda p: p.stat().st_mtime)
    if not all_master_files:
        print(f"ERROR: No master xlsx files found in {MASTER_OUTPUT_DIR}")
        sys.exit(1)
    master_files = [all_master_files[-1]]  # only the latest scrape run

    # Step 1: Load taxonomy
    print("Loading taxonomy...")
    alias_map = load_taxonomy(TAXONOMY_JSON)
    print(f"  {len(alias_map)} aliases")

    # Step 2: Read raw jobs (deduped by job_id)
    print(f"\nReading latest master file: {master_files[0].name}")
    raw_jobs = read_all_jobs(master_files)
    print(f"  {len(raw_jobs)} unique jobs loaded")

    # Step 3: Preprocess
    print("\nPreprocessing...")
    jobs, stats = preprocess(raw_jobs)
    print(f"  {stats['dropped_quality']} dropped — JD too short (<150 chars)")
    print(f"  {stats['dropped_staleness']} dropped — posted >50 days before most recent")
    print(f"  {stats['output']} jobs remaining")

    # Step 4: Tag with LLM (5-provider fallback chain — see skill_tagger.py)
    print("\nTagging skills with LLM (cached)...")
    api_keys = {
        "LM_STUDIO_TAGGER_MODEL": lm_studio_tagger_model,
        "GEMINI_API_KEY":         os.getenv("GEMINI_API_KEY", ""),
        "CEREBRAS_API_KEY":       os.getenv("CEREBRAS_API_KEY", ""),
        "GROQ_API_KEY":           groq_api_key,
        "SAMBANOVA_API_KEY":      os.getenv("SAMBANOVA_API_KEY", ""),
        "OPENROUTER_API_KEY":     os.getenv("OPENROUTER_API_KEY", ""),
    }
    from app.services.skill_tagger import tag_jobs_with_llm
    tag_cache, _ = tag_jobs_with_llm(jobs, alias_map, api_keys, verbose=True)

    # Step 5: Drop jobs with zero required skills AND no raw skill fallback.
    # Keep jobs that have raw skills — csv_importer will use them as fallback.
    before = len(jobs)
    jobs = [
        j for j in jobs
        if tag_cache.get(str(j["job_id"]), {}).get("required_skills")
        or str(j.get("skills_required_raw") or "").strip()
    ]
    dropped_zero = before - len(jobs)
    if dropped_zero:
        print(f"  {dropped_zero} dropped — zero LM skills AND no raw skill fallback")

    # Step 6: JD field enrichment (seniority, work mode, industry, degree, etc.)
    lm_studio_extractor = (
        os.getenv("LM_STUDIO_EXTRACTOR_MODEL", "")
        or lm_studio_tagger_model  # fall back to tagger model if extractor not set
    )
    enrichment_cache: dict = {}
    if lm_studio_extractor:
        print("\nEnriching structured fields from JD text (cached)...")
        from app.services.jd_enricher import enrich_jobs
        enrichment_cache = enrich_jobs(jobs, model=lm_studio_extractor, verbose=True)
    else:
        print("\nSkipping JD enrichment — no LM Studio model configured.")

    # Step 7: Write enhanced xlsx
    output_path = ENHANCED_OUTPUT_DIR / f"ENHANCED_JOBS_{date.today().strftime('%Y%m%d')}.xlsx"
    write_enhanced_xlsx(jobs, tag_cache, output_path, enrichment_cache)

    print(f"\nDone. {len(jobs)} jobs written to:")
    print(f"  {output_path}")
    print("\nReview the enhanced xlsx, then run csv_importer.py to push to Supabase.")


if __name__ == "__main__":
    main()
